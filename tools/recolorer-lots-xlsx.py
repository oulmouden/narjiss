# -*- coding: utf-8 -*-
"""Aligne les classeurs de lots sur la palette de shared/statuts-lots.css.

Deux gestes :
  1. les aplats deja poses (onglets Lots et Synthese du classeur de demo)
     passent aux nouvelles teintes ;
  2. une mise en forme conditionnelle est posee sur la colonne « statut »,
     pour que les lignes saisies plus tard se colorent seules — le modele
     vierge n'avait aucune couleur, le bureau de vente les posait a la main.
"""
import io
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

# Statut -> (aplat, encre). Aplats : versions claires des teintes du site,
# lisibles derriere du texte ; encres : exactement les valeurs CSS.
PALETTE = {
    'Disponible': ('E4F1EA', '1F6F55'),
    'Optionne':   ('FBEEDA', '9E6300'),
    'Reserve':    ('E3EDF9', '1C5FA8'),
    'Vendu':      ('ECEEF1', '5A6272'),
    'Bloque':     ('F7DEDE', 'B02A2A'),
}

# Anciens aplats -> statut, pour retrouver ce que chaque cellule voulait dire.
ANCIENS = {
    'D9EAD3': 'Disponible',
    'FFF2CC': 'Optionne',
    'FCE5CD': 'Reserve',
    'E6E6E6': 'Vendu',
    'F4CCCC': 'Bloque',
}

COLONNE_STATUT = 'Q'


def code(cellule):
    f = cellule.fill
    if not f or f.fill_type != 'solid':
        return None
    rgb = getattr(f.start_color, 'rgb', None)
    if not isinstance(rgb, str):
        return None
    return rgb[-6:].upper()


def recolorer(ws):
    n = 0
    for ligne in ws.iter_rows():
        for c in ligne:
            statut = ANCIENS.get(code(c))
            if not statut:
                continue
            aplat, encre = PALETTE[statut]
            c.fill = PatternFill('solid', start_color=aplat, end_color=aplat)
            base = c.font
            c.font = Font(name=base.name, size=base.size, bold=base.bold,
                          italic=base.italic, color=encre)
            n += 1
    return n


def poser_regles(ws, derniere=1000):
    """Colore la cellule de statut selon son texte, pour toute ligne a venir."""
    plage = '%s2:%s%d' % (COLONNE_STATUT, COLONNE_STATUT, derniere)
    for statut, (aplat, encre) in PALETTE.items():
        style = DifferentialStyle(
            fill=PatternFill(bgColor=aplat),
            font=Font(color=encre, bold=True),
        )
        regle = Rule(type='containsText', operator='containsText', text=statut,
                     dxf=style, stopIfTrue=False)
        regle.formula = ['NOT(ISERROR(SEARCH("%s",%s2)))' % (statut, COLONNE_STATUT)]
        ws.conditional_formatting.add(plage, regle)


def traiter(chemin):
    wb = load_workbook(chemin)
    total = 0
    # Seuls les onglets qui parlent de statuts. L'onglet Notice utilise le meme
    # jaune (FFF2CC) pour surligner ses questions au client : le repeindre en
    # ambre << optionne >> lui ferait dire ce qu'il ne dit pas.
    for nom in ('Lots', 'Synthese'):
        if nom in wb.sheetnames:
            total += recolorer(wb[nom])
    if 'Lots' in wb.sheetnames:
        poser_regles(wb['Lots'])
    wb.save(chemin)
    print('%s : %d cellules recolorees, regles posees sur Lots!%s' % (chemin, total, COLONNE_STATUT))


if __name__ == '__main__':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    for chemin in sys.argv[1:]:
        traiter(chemin)
